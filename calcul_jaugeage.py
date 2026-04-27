"""
calcul_jaugeage.py
Module de calcul de jaugeage pour citerne cylindrique verticale.

Géométrie :
  - Fond bas  : PLAT  → V(h) = π × R² × h  (formule exacte)
  - Corps     : cylindrique
  - Fond haut : BOMBÉ elliptique
                V(h) = π × R² × h² / h_dome × (1 - h / (3 × h_dome))
"""

import math
import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Classe principale
# ---------------------------------------------------------------------------

class CiterneVertical:
    """
    Citerne cylindrique verticale :
      - Fond bas  : PLAT  (h = 0 → HT)
      - Corps     : cylindrique
      - Fond haut : elliptique bombé (h = HT → HF)

    HT  = hauteur du corps cylindrique (depuis le fond plat jusqu'à la tangente du dome)
    HF  = hauteur totale (fond plat → sommet du dome)
    h_dome = HF - HT = hauteur du fond bombé supérieur
    """

    def __init__(self, diametre_mm: float, HF_mm: float, HT_mm: float,
                 appellation: str = ""):
        self.appellation = appellation
        self.D = float(diametre_mm)
        self.R = self.D / 2
        self.HF = float(HF_mm)
        self.HT = float(HT_mm)
        self.h_dome = self.HF - self.HT   # hauteur du fond bombé haut uniquement

    # --- volume du fond bombé supérieur ---

    def _V_dome(self, h: float) -> float:
        """
        Volume (mm³) du fond elliptique supérieur rempli jusqu'à h
        depuis le début du dome (0 ≤ h ≤ h_dome).
        Formule : π × R² × h² / h_dome × (1 - h / (3 × h_dome))
        """
        h = max(0.0, min(h, self.h_dome))
        if h == 0.0:
            return 0.0
        return math.pi * self.R**2 * h**2 / self.h_dome * (1.0 - h / (3.0 * self.h_dome))

    def _V_dome_complet(self) -> float:
        """Volume total du fond bombé = 2/3 × π × R² × h_dome."""
        return (2.0 / 3.0) * math.pi * self.R**2 * self.h_dome

    # --- volume total à hauteur h ---

    def volume_mm3(self, h: float) -> float:
        """
        Volume cumulé (mm³) depuis le fond plat jusqu'à la hauteur h (mm).

        Zone corps  (0 ≤ h ≤ HT)  : V = π × R² × h  (fond plat → formule exacte)
        Zone dome   (HT < h ≤ HF) : V = V_corps_complet + V_dome(h - HT)
        """
        h = max(0.0, min(float(h), self.HF))

        if h <= self.HT:
            # Fond plat + corps cylindrique
            return math.pi * self.R**2 * h
        else:
            # Corps complet + portion du dome supérieur
            V_corps = math.pi * self.R**2 * self.HT
            return V_corps + self._V_dome(h - self.HT)

    def volume_L(self, h: float) -> float:
        return self.volume_mm3(h) / 1_000_000.0

    def volume_m3(self, h: float) -> float:
        return self.volume_mm3(h) / 1_000_000_000.0

    def delta_L_par_mm(self, h: float) -> float:
        if h < 1:
            return self.volume_L(1.0)
        return self.volume_L(h) - self.volume_L(h - 1.0)

    def zone(self, h: float) -> str:
        if h <= self.HT:
            return "Corps (fond plat)"
        else:
            return "Fond bombé haut"

    # --- construction du tableau ---

    def build_jaugeage(self, H_mort: float) -> pd.DataFrame:
        """
        Retourne un DataFrame avec une ligne par mm de 0 à HF.
        H_mort : hauteur du volume mort (mm) — en dessous, Vol. utilisable = 0.
        """
        V_mort = self.volume_L(H_mort)
        rows = []
        for h in range(0, int(self.HF) + 1):
            vL = self.volume_L(h)
            vol_utile = 0.0 if h <= H_mort else round(vL - V_mort, 2)
            rows.append({
                "Hauteur (mm)": h,
                "Volume (L)": round(vL, 2),
                "Vol. utilisable (L)": vol_utile,
                "Volume (m³)": round(self.volume_m3(h), 4),
                "ΔV (L/mm)": round(self.delta_L_par_mm(h), 2),
                "Zone": self.zone(h),
            })
        return pd.DataFrame(rows)

    # --- résumé ---

    def resume(self, H_mort: float, H_aspiration: float) -> dict:
        return {
            "Volume corps cylindrique (L)": math.pi * self.R**2 * self.HT / 1e6,
            "Volume fond bombé haut (L)":  self._V_dome_complet() / 1e6,
            "Volume total (L)":            self.volume_L(self.HF),
            "Volume total (m³)":           self.volume_m3(self.HF),
            "Volume mort (L)":             self.volume_L(H_mort),
            "Volume à aspiration (L)":     self.volume_L(H_aspiration),
            "Volume utile (L)":            self.volume_L(self.HF) - self.volume_L(H_mort),
        }


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

def hauteur_pour_volume(citerne: CiterneVertical, volume: float, unite: str = "L",
                        tolerance_mm: float = 0.1) -> dict:
    """
    Recherche inverse : retourne la hauteur (mm) correspondant au volume donné.

    unite   : "L" (litres), "m3" (mètres cubes) ou "mm3" (millimètres cubes)
    Méthode : recherche binaire sur volume_mm3(h) qui est strictement croissante.
    Retourne un dict avec hauteur, volume exact à cette hauteur, et zone.
    """
    if unite == "L":
        V_mm3 = volume * 1_000_000.0
    elif unite == "m3":
        V_mm3 = volume * 1_000_000_000.0
    else:
        V_mm3 = float(volume)

    V_min = citerne.volume_mm3(0)
    V_max = citerne.volume_mm3(citerne.HF)

    if V_mm3 <= V_min:
        h = 0.0
    elif V_mm3 >= V_max:
        h = citerne.HF
    else:
        lo, hi = 0.0, citerne.HF
        while hi - lo > tolerance_mm:
            mid = (lo + hi) / 2.0
            if citerne.volume_mm3(mid) < V_mm3:
                lo = mid
            else:
                hi = mid
        h = (lo + hi) / 2.0

    return {
        "hauteur_mm": round(h, 1),
        "volume_L":   round(citerne.volume_L(h), 2),
        "volume_m3":  round(citerne.volume_m3(h), 4),
        "volume_mm3": round(citerne.volume_mm3(h), 0),
        "zone":       citerne.zone(h),
    }


def export_excel(
    citerne: CiterneVertical,
    df: pd.DataFrame,
    H_mort: float,
    H_aspiration: float,
    output=None,
) -> str | io.BytesIO:
    """
    Génère le fichier Excel de jaugeage.

    output : chemin de fichier (str) ou None → retourne un BytesIO.
    """

    # ---- couleurs ----
    C_BLEU_FONCE  = "1F4E79"
    C_BLEU_CLAIR  = "BDD7EE"
    C_MORT        = "FFCCCC"   # rouge clair = volume mort (inaccessible)
    C_MORT_BORD   = "FF4C4C"   # rouge vif = limite exacte volume mort
    C_VERT        = "C6EFCE"
    C_ROUGE_FOND  = "C00000"
    C_BLANC       = "FFFFFF"
    C_GRIS        = "F2F2F2"

    def fill(c):
        return PatternFill("solid", fgColor=c)

    thin = Side(style="thin")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ==================================================================
    # FEUILLE 1 — Tableau de jaugeage
    # ==================================================================
    ws = wb.active
    ws.title = "Jaugeage"

    # Titre
    ws["A1"] = f"TABLEAU DE JAUGEAGE — {citerne.appellation}"
    ws["A1"].font = Font(bold=True, size=14, color=C_BLANC)
    ws["A1"].fill = fill(C_BLEU_FONCE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 32

    # Sous-titre
    ws["A2"] = (
        f"φ={citerne.D:,.0f} mm  |  HF={citerne.HF:,.0f} mm  |  "
        f"HT={citerne.HT:,.0f} mm  |  h_fond={citerne.h_fond:.0f} mm  |  "
        f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].fill = fill(C_BLEU_CLAIR)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 16

    # Légende couleurs
    ws["A3"] = f"Rouge = VOLUME MORT (≤ {H_mort:.0f} mm) — NE PEUT PAS ÊTRE ASPIRÉ — Vol. utilisable = 0"
    ws["A3"].fill = fill(C_MORT)
    ws["A3"].font = Font(size=9, bold=True, color="C00000")
    ws.merge_cells("A3:C3")
    ws["D3"] = "Vert = Aspiration"
    ws["D3"].fill = fill(C_VERT)
    ws["D3"].font = Font(size=9)
    ws["E3"] = "Bleu = Fonds elliptiques"
    ws["E3"].fill = fill(C_BLEU_CLAIR)
    ws["E3"].font = Font(size=9)
    ws["F3"] = "Blanc/Gris = Corps cylindrique"
    ws["F3"].font = Font(size=9)
    ws.row_dimensions[3].height = 16

    # En-têtes colonnes
    headers = list(df.columns)
    nb_cols = len(headers)
    last_col_letter = get_column_letter(nb_cols)
    for col_i, h_txt in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_i, value=h_txt)
        cell.font = Font(bold=True, color=C_BLANC, size=11)
        cell.fill = fill(C_BLEU_FONCE)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bord
    ws.row_dimensions[4].height = 22

    # Merge titre et sous-titre sur nb_cols colonnes
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.merge_cells(f"A2:{last_col_letter}2")

    # Données
    for idx, row_data in df.iterrows():
        r = idx + 5
        h_val = int(row_data["Hauteur (mm)"])

        if h_val < H_mort:
            bg = fill(C_MORT)           # rouge clair = zone morte inaccessible
        elif h_val == H_mort:
            bg = fill(C_MORT_BORD)      # rouge vif = limite exacte du volume mort
        elif h_val == H_aspiration:
            bg = fill(C_VERT)
        elif row_data["Zone"] != "Corps":
            bg = fill(C_BLEU_CLAIR)
        elif idx % 2 == 0:
            bg = fill(C_GRIS)
        else:
            bg = fill(C_BLANC)

        for col_i, val in enumerate(row_data.values, 1):
            cell = ws.cell(row=r, column=col_i, value=val)
            cell.fill = bg
            cell.border = bord
            # Colonne "Vol. utilisable" : mettre en gras les 0 pour signaler le mort
            if col_i == 3 and h_val <= H_mort:
                cell.font = Font(bold=True, color="C00000")
            cell.alignment = Alignment(
                horizontal="right" if col_i < nb_cols else "center"
            )

    # Largeurs
    for col_letter, w in zip("ABCDEF", [15, 18, 20, 15, 14, 18]):
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_col_letter}{4 + len(df)}"

    # ==================================================================
    # FEUILLE 2 — Paramètres
    # ==================================================================
    ws2 = wb.create_sheet("Paramètres")

    ws2["A1"] = "PARAMÈTRES DE LA CITERNE"
    ws2["A1"].font = Font(bold=True, size=13, color=C_BLANC)
    ws2["A1"].fill = fill(C_BLEU_FONCE)
    ws2.merge_cells("A1:B1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    params = [
        ("Appellation",                     citerne.appellation),
        ("Diamètre intérieur (D)",          f"{citerne.D:,.0f} mm"),
        ("Rayon (R)",                       f"{citerne.R:,.0f} mm"),
        ("Circonférence",                   f"{math.pi * citerne.D:,.1f} mm"),
        ("Hauteur totale HF",               f"{citerne.HF:,.0f} mm"),
        ("Hauteur corps cylindrique HT",    f"{citerne.HT:,.0f} mm"),
        ("Hauteur fond bombé supérieur",    f"{citerne.h_dome:.1f} mm"),
        ("Fond inférieur",                  "PLAT"),
        ("", ""),
        ("Volume corps cylindrique",        f"{math.pi*citerne.R**2*citerne.HT/1e6:,.1f} L"),
        ("Volume fond bombé haut (complet)",f"{citerne._V_dome_complet()/1e6:,.1f} L"),
        ("VOLUME TOTAL CALCULÉ",            f"{citerne.volume_L(citerne.HF):,.1f} L  /  {citerne.volume_m3(citerne.HF):,.2f} m³"),
        ("Volume nominal (référence)",      "1 200 000 L"),
        ("", ""),
        ("Hauteur volume mort",             f"{H_mort:.0f} mm"),
        ("Volume mort (= π×R²×H_mort)",    f"{citerne.volume_L(H_mort):,.1f} L"),
        ("Hauteur aspiration",              f"{H_aspiration:.0f} mm"),
        ("Volume à l'aspiration",           f"{citerne.volume_L(H_aspiration):,.1f} L"),
        ("Volume utile (aspir. → plein)",   f"{citerne.volume_L(citerne.HF)-citerne.volume_L(H_mort):,.1f} L"),
        ("ΔV par mm (corps cylindrique)",   f"{citerne.delta_L_par_mm(citerne.HT/2):,.2f} L/mm"),
    ]

    for row_i, (label, value) in enumerate(params, 2):
        cl = ws2.cell(row=row_i, column=1, value=label)
        cv = ws2.cell(row=row_i, column=2, value=value)
        if label == "VOLUME TOTAL CALCULÉ":
            for c in (cl, cv):
                c.font = Font(bold=True, color=C_BLANC)
                c.fill = fill(C_BLEU_FONCE)
        elif label:
            cl.font = Font(bold=True)

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 35

    # ==================================================================
    # FEUILLE 3 — Points clés
    # ==================================================================
    ws3 = wb.create_sheet("Points Clés")

    ws3["A1"] = "POINTS CLÉS DU JAUGEAGE"
    ws3["A1"].font = Font(bold=True, size=13, color=C_BLANC)
    ws3["A1"].fill = fill(C_BLEU_FONCE)
    ws3.merge_cells("A1:D1")
    ws3["A1"].alignment = Alignment(horizontal="center")

    kp_headers = ["Description", "Hauteur (mm)", "Volume (L)", "Volume (m³)"]
    for col_i, h_txt in enumerate(kp_headers, 1):
        c = ws3.cell(row=2, column=col_i, value=h_txt)
        c.font = Font(bold=True, color=C_BLANC)
        c.fill = fill(C_BLEU_FONCE)
        c.border = bord

    key_points = [
        ("Fond plat vide (h = 0)",                0,           fill(C_BLANC)),
        (f"Volume mort  (H={H_mort:.0f} mm)",     H_mort,      fill(C_MORT_BORD)),
        (f"Aspiration   (H={H_aspiration:.0f} mm)", H_aspiration, fill(C_VERT)),
        ("Mi-hauteur corps",                       citerne.HT / 2,  fill(C_BLANC)),
        ("Début fond bombé (H = HT)",              citerne.HT,  fill(C_BLEU_CLAIR)),
        ("PLEIN (100 %)",                          citerne.HF,  fill(C_ROUGE_FOND)),
    ]

    for row_i, (desc, h_val, bg) in enumerate(key_points, 3):
        vL  = citerne.volume_L(h_val)
        vm3 = citerne.volume_m3(h_val)
        vals = [desc, int(h_val), round(vL, 1), round(vm3, 4)]
        is_plein = desc.startswith("PLEIN")
        for col_i, val in enumerate(vals, 1):
            c = ws3.cell(row=row_i, column=col_i, value=val)
            c.fill = bg
            c.border = bord
            if is_plein:
                c.font = Font(bold=True, color=C_BLANC)

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 15

    # ==================================================================
    # Sauvegarde
    # ==================================================================
    if output is None:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    else:
        wb.save(output)
        return output


# ---------------------------------------------------------------------------
# Point d'entrée CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    # Paramètres de la citerne — modifiables ici ou via l'interface Streamlit
    citerne = CiterneVertical(
        diametre_mm=12_000,
        HF_mm=12_080,
        HT_mm=10_630,
        appellation="BAC R6",
    )

    # Ces valeurs DOIVENT être saisies — aucune valeur par défaut imposée
    try:
        H_MORT       = float(sys.argv[1]) if len(sys.argv) > 1 else float(input("Hauteur volume mort (mm) : "))
        H_ASPIRATION = float(sys.argv[2]) if len(sys.argv) > 2 else float(input("Hauteur aspiration  (mm) : "))
    except ValueError:
        print("Erreur : entrez des valeurs numériques.")
        sys.exit(1)

    print("=" * 55)
    print(f"  JAUGEAGE — {citerne.appellation}")
    print("=" * 55)
    print(f"  Diamètre     : {citerne.D:,.0f} mm")
    print(f"  HF (fond/fond): {citerne.HF:,.0f} mm")
    print(f"  HT (corps)   : {citerne.HT:,.0f} mm")
    print(f"  h_fond       : {citerne.h_fond:.1f} mm")
    print(f"  Volume total : {citerne.volume_L(citerne.HF):,.1f} L")
    print(f"  Volume mort  : {citerne.volume_L(H_MORT):,.1f} L  (H={H_MORT} mm)")
    print(f"  Vol. aspir.  : {citerne.volume_L(H_ASPIRATION):,.1f} L  (H={H_ASPIRATION} mm)")
    print("-" * 55)

    print("Construction du tableau (12 081 lignes)…", end=" ", flush=True)
    df = citerne.build_jaugeage(H_mort=H_MORT)
    print("OK")

    filename = "Jaugeage_BAC_R6.xlsx"
    export_excel(citerne, df, H_mort=H_MORT, H_aspiration=H_ASPIRATION, output=filename)
    print(f"Fichier généré : {filename}")
